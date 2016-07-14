# -*- coding: utf-8 -*-

import os
import sys
import shutil
import traceback
import simplejson
import hashlib
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import datetime
import smtplib
from subprocess import Popen, PIPE, STDOUT
import bagit





startTime = time.time()
startTimeReadable = str(datetime.datetime.now().isoformat())
print startTimeReadable

#start log
startLog = open("log.txt", "a")
logText = "\n****************************************************************************************************************\n"
logText = logText + "Crawl started " + startTimeReadable
startLog.write(logText)
startLog.close()

try:

	if os.name == "nt":
		#triageDir = "\\\\LINCOLN\\Library\\UA200"
		triageDir = "\\\\romeo\\Collect\\spe\\Greg\\ua200_testing"
		presDir = "\\\\LINCOLN\\Masters\\Special Collections\\accessions"
		createSIPDir = ""
	else:
		#triageDir = "/media/bcadmin/Lincoln/Library/UA200"
		triageDir = "/media/bcadmin/Collect/spe/Greg/ua200_testing"
		presDir = "/media/bcadmin/Lincoln/Special Collections/accessions"
		createSIPDir = ""
		
	#from http://stackoverflow.com/questions/14996453/python-libraries-to-calculate-human-readable-filesize-from-bytes
	suffixes = ['B', 'KB', 'MB', 'GB', 'TB', 'PB']
	def humansize(nbytes):
		if nbytes == 0: return '0 B'
		i = 0
		while nbytes >= 1024 and i < len(suffixes)-1:
			nbytes /= 1024.
			i += 1
		f = ('%.2f' % nbytes).rstrip('0').rstrip('.')
		return '%s %s' % (f, suffixes[i])
		
	hashDir = os.path.join(presDir, "hashDir", "ua200Hash")
	if not os.path.isdir(hashDir):
		os.makedirs(hashDir)
	logDir = os.path.join(triageDir, "Logs")

	#hash function
	#from http://stackoverflow.com/questions/3431825/generating-an-md5-checksum-of-a-file
	def md5(fname):
		hash_md5 = hashlib.md5()
		with open(fname, "rb") as f:
			for chunk in iter(lambda: f.read(4096), b""):
				hash_md5.update(chunk)
		return hash_md5.hexdigest()

	for folder in os.listdir(triageDir):
		print folder
		if folder.lower() == "requests" or folder.lower() == "logs" or folder.lower() == "thumbs.db":
			pass
		elif not os.path.isdir(os.path.join(triageDir, folder)):
			pass
		elif folder.lower() == "councils":
			for subfolder in os.listdir(os.path.join(triageDir, folder)):
				if os.path.isdir(os.path.join(triageDir, folder, subfolder)):
					print "	reading: " + subfolder
					
					#get log for series
					accessionFile = os.path.join(logDir, folder, subfolder + ".xlsx")
					if not os.path.isdir(os.path.join(logDir, folder)):
						os.makedirs(os.path.join(logDir, folder))
					if not os.path.isfile(accessionFile):
						print "no log found for " + subfolder
						print "creating empty log"
						seriesLogBook = Workbook()
					else:
						seriesLogBook = load_workbook(accessionFile, False)
					seriesLog = seriesLogBook.active
									
					#get hashIndex for series
					if not os.path.isdir(os.path.join(hashDir, folder)):
						os.makedirs(os.path.join(hashDir, folder))
					if not os.path.isfile(os.path.join(hashDir, folder, subfolder + "Hash.json")):
						print "no hashIndex found for " + subfolder
						print "creating empty hashIndex"
						newIndex = open(os.path.join(hashDir, folder, subfolder + "Hash.json"), "w")
						newIndex.write("{}")
						newIndex.close()
									
					with open(os.path.join(hashDir, folder, subfolder + "Hash.json"), "r") as hashFile:
						hashIndex = simplejson.loads(hashFile.read())
						hashFile.close()
					
					for series, paths, files in os.walk(os.path.join(triageDir, folder, subfolder).decode(sys.getfilesystemencoding())):
						for file in files:
							
							filePath = os.path.join(series, file)
							print "found " + file.encode(sys.getfilesystemencoding())
							if os.name == "nt":
								subPath = filePath.split("\\" + subfolder + "\\")[1]
							else:
								subPath = filePath.split("/" + subfolder + "/")[1]
							subPath = os.path.dirname(subPath)
							
							hash = str(md5(filePath))
							if hash in hashIndex.values():
								print "hash found, file was previously transferred"
								aquTime = time.time()
								aquTimeReadable = str(time.strftime("%Y-%m-%d %H:%M:%S"))
								seriesLog.append([subPath, file, aquTimeReadable, "DUPLICATE", aquTime])
							else:
																
								if not os.path.isdir(os.path.join(triageDir, "ua200", folder, subfolder, subPath)):
									os.makedirs(os.path.join(triageDir, "ua200", folder, subfolder, subPath))
								shutil.move(filePath, os.path.join(triageDir, "ua200", folder, subfolder, subPath))
								aquTime = time.time()
								aquTimeReadable = str(time.strftime("%Y-%m-%d %H:%M:%S"))
								hashIndex.update({file: hash})
								seriesLog.append([subPath, file, aquTimeReadable, hash, aquTime])
								
					with open(os.path.join(hashDir, folder, subfolder + "Hash.json"), "w") as hashFile:
						simplejson.dump(hashIndex, hashFile)
						hashFile.close()
						
					seriesLogBook.save(accessionFile)
					#remove empty directories
					print "clearing " + subfolder
					emptyCheck = True
					for folderRoot, subFolders, subFiles in os.walk(subfolder):
						for fileEx in subfiles:
							emptyCheck = True
							raise ValueError("Error: files still present in " + folder)
					if emptyCheck == True:
						print subfolder + " is empty"
						folderName = folder
						subfolderName = subfolder
						shutil.rmtree(os.path.join(triageDir, folder, subfolder))
						os.makedirs(os.path.join(triageDir, folderName, subfolderName))
											
					
					
		else:
			print "reading " + folder
			
			#get log for series
			accessionFile = os.path.join(logDir, folder + ".xlsx")
			if not os.path.isfile(accessionFile):
				print "no log found for " + folder
				print "creating empty log"
				seriesLogBook = Workbook()
			else:
				seriesLogBook = load_workbook(accessionFile, False)
			seriesLog = seriesLogBook.active
							
			#get hashIndex for series
			if not os.path.isfile(os.path.join(hashDir, folder + "Hash.json")):
				print "no hashIndex found for " + folder
				print "creating empty hashIndex"
				newIndex = open(os.path.join(hashDir, folder + "Hash.json"), "w")
				newIndex.write("{}")
				newIndex.close()
							
			with open(os.path.join(hashDir, folder + "Hash.json"), "r") as hashFile:
				hashIndex = simplejson.loads(hashFile.read())
				hashFile.close()
			
			for series, paths, files in os.walk(os.path.join(triageDir, folder).decode(sys.getfilesystemencoding())):
				for file in files:
					filePath = os.path.join(series, file)
					print "found " + file.encode(sys.getfilesystemencoding())
					if os.name == "nt":
						subPath = filePath.split("\\" + folder + "\\")[1]
					else:
						subPath = filePath.split("/" + folder + "/")[1]
					subPath = os.path.dirname(subPath)
					
					hash = str(md5(filePath))
					if hash in hashIndex.values():
						print "hash found, file was previously transferred"
						aquTime = time.time()
						aquTimeReadable = str(time.strftime("%Y-%m-%d %H:%M:%S"))
						seriesLog.append([subPath, file, aquTimeReadable, "DUPLICATE", aquTime])
					else:
						
						if not os.path.isdir(os.path.join(triageDir, "ua200", folder, subPath)):
							os.makedirs(os.path.join(triageDir, "ua200", folder, subPath))
						shutil.move(filePath, os.path.join(triageDir, "ua200", folder, subPath))
						aquTime = time.time()
						aquTimeReadable = str(datetime.datetime.now().isoformat()).split(".")[0].replace("T", " ")
						hashIndex.update({file: hash})
						seriesLog.append([subPath, file, aquTimeReadable, hash, aquTime])
						
			with open(os.path.join(hashDir, folder + "Hash.json"), "w") as hashFile:
				simplejson.dump(hashIndex, hashFile)
				hashFile.close()
				
			seriesLogBook.save(accessionFile)
			#remove empty directories
			print "clearing " + folder
			emptyCheck = True
			for folderRoot, subFolders, subFiles in os.walk(folder):
				for fileEx in subfiles:
					emptyCheck = True
					raise ValueError("Error: files still present in " + folder)
			if emptyCheck == True:
				print folder + " is empty"
				folderName = folder
				shutil.rmtree(os.path.join(triageDir, folder))
				os.mkdir(os.path.join(triageDir, folderName))
		
	#get file and size count	
	fileCount = 0
	totalSize = 0
	for root, dirs, files in os.walk(os.path.join(triageDir, "ua200")):
		fileCount += len(files)
		for f in files:
			fp = os.path.join(root, f)
			totalSize += os.path.getsize(fp)
	readableSize = humansize(totalSize)
	
	if fileCount > 0:
	
		#createSIP
		print "bagging accession"
		if os.name == "nt":
			sipCmd = "python C:\\Projects\\createsip\\createsip.py " + os.path.join(triageDir, "ua200")
		else:
			sipCmd = "sudo python /home/bcadmin/Projects/createSIP/createSIP.py " + os.path.join(triageDir, "ua200")
		createSIP = Popen(sipCmd, stdout=PIPE, stdin=PIPE, stderr=STDOUT)
		stdout, stderr = createSIP.communicate("ua200\nElisa Lopez\nRecords from the University Senate\n\nua200.py crawler\n\nUniversity Senate\nSecretary, Manages Senate Records\nemlopez@albany.edu\nUNH 302\n\n\n\n\nY\n")
		SIP = str(stdout).split('|||')[1].strip()
		
		#validate bag
		print "validating bag"
		bagValid = False
		bagCheck = bagit.Bag(SIP)
		if bagCheck.is_valid():
			bagValid = True
			print "bag is valid!"
		else:
			bagValid = False
			print "bag failed to validate"
			print os.name
			
			
		#make copy of logs
		print "making copies of logs"
		for logFile in os.listdir(logDir.decode(sys.getfilesystemencoding())):
			print logFile
			if os.path.isfile(os.path.join(logDir, logFile)):
				logCopy = os.path.join(presDir, "crawlerLogs", os.path.basename(logFile))
				shutil.copy(os.path.join(logDir, logFile), logCopy)
			else:
				if not os.path.isdir(os.path.join(presDir, "crawlerLogs", logFile)):
					os.makedirs(os.path.join(presDir, "crawlerLogs", logFile))
				for logSubFile in os.listdir(os.path.join(logDir, logFile).decode(sys.getfilesystemencoding())):
					logCopy = os.path.join(presDir, "crawlerLogs", logFile, os.path.basename(logSubFile))
					shutil.copy(os.path.join(logDir, logFile, logSubFile), logCopy)
					
	else:
		print "no new files found"
			
	finalTime = time.time() - startTime
	print "Total Time: " + str(finalTime) + " seconds, " + str(finalTime/60) + " minutes, " + str(finalTime/3600) + " hours"
	finalTimeFile = open("log.txt", "a")
	logText = "\nSuccessful Crawl ran " + str(time.strftime("%Y-%m-%d %H:%M:%S"))
	if bagValid == True:
		logText = logText + "\nBag is Valid!"
	else:
		logText = logText + "\nBag failed to validate."
	logText = logText + "\nProcess took " + str(finalTime) + " seconds or " + str(finalTime/60) + " minutes or " + str(finalTime/3600) + " hours"
	logText = logText + "\n" + str(fileCount) + " files transferred."
	logText = logText + "\n" + str(totalSize) + " bytes or " + str(readableSize) + " transferred."
	finalTimeFile.write(logText)
	finalTimeFile.close()

	sender = 'UAlbanyArchivesNotify@gmail.com'
	receivers = ['gwiedeman@albany.edu']
	subject = "UA200 Crawler Success"
	body = logText
	message = 'Subject: %s\n\n%s' % (subject, body)
	smtpObj = smtplib.SMTP(host='smtp.gmail.com', port=587)
	smtpObj.ehlo()
	smtpObj.starttls()
	smtpObj.ehlo()
	keyFile = open("pw.txt", "r")
	lines = keyFile.readlines()
	emailPW = lines[0]
	keyFile.close()
	smtpObj.login('UAlbanyArchivesNotify', emailPW)
	smtpObj.sendmail(sender, receivers, message)
	smtpObj.quit()

except:
	exceptMsg = str(traceback.format_exc())

	updateLog = open("log.txt", "a")
	logText = "\nCrawl failed at " + str(time.strftime("%Y-%m-%d %H:%M:%S"))
	updateLog.write(logText)
	updateLog.close()

	finalTime = time.time() - startTime
	print "Total Time: " + str(finalTime) + " seconds, " + str(finalTime/60) + " minutes, " + str(finalTime/3600) + " hours"
	print exceptMsg
	errorLog = open("errorLog.txt", "a")
	errorText = "***********************************************************************************\n" + str(time.strftime("%Y-%m-%d %H:%M:%S")) + "\n" + str(finalTime) + " seconds\n" + str(finalTime/60) + " minutes\n" + str(finalTime/3600) + " hours" + "\nTraceback:\n" + exceptMsg
	errorLog.write(errorText)
	errorLog.close()
	
	sender = 'UAlbanyArchivesNotify@gmail.com'
	receivers = ['gwiedeman@albany.edu']
	subject = "UA200 Crawler Error"

	body = "ERROR: " + logText + "\n\n" + exceptMsg
	message = 'Subject: %s\n\n%s' % (subject, body)
	smtpObj = smtplib.SMTP(host='smtp.gmail.com', port=587)
	smtpObj.ehlo()
	smtpObj.starttls()
	smtpObj.ehlo()
	keyFile = open("pw.txt", "r")
	lines = keyFile.readlines()
	emailPW = lines[0]
	keyFile.close()
	smtpObj.login('UAlbanyArchivesNotify', emailPW)
	smtpObj.sendmail(sender, receivers, message)
	smtpObj.quit()